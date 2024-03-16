import openpyxl


def test_ddt():
    file = "C:\\Users\\esita\\PycharmProjects\\selenium_practice\\ddt\\testdata_ddt.xlsx"
    #file="C:\\Users\\esita\\PycharmProjects\\selenium_practice\\ddt\\vwo_test_case.xlsx"
    workbook = openpyxl.load_workbook(file)
    sheet = workbook["testdata_ddt"]
    #sheet=workbook["Project1"]
    rows = sheet.max_row
    cols = sheet.max_column
    for r in range(1, rows + 1):
        for c in range(1, cols + 1):
            print(sheet.cell(r, c).value, end='     ')
        print()
